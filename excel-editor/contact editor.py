import os 
import openpyxl


try:
    
    folder = r'E:\works\python'
    wb = openpyxl.load_workbook('contactlist.xlsx')
    sheet = wb['contacts'] 

    for i in range(5):
        
        cell = sheet.cell(row=i,column=2)
        x = cell.value
        print(x)

        wb.save('contactlist.xlsx')

    '''
    cell = sheet.cell(row=1, column=1)

    cell.value = "New Value"
    '''



except:
    print("error")
